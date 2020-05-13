import queue
from chrom import *
from random import *
from botClass import *
from threading import Thread
from statistics import *
from openpyxl import Workbook
import openpyxl

#IF YOU WANT TO TRAIN, MAKE SURE YOU SCROLL DOWN AND UNCOMMENT ONE OF THE TRAINING CODES

class QGA:

    def __init__(self, bit, len, rMutate, typ, wList, limit):
        self.len = len
        self.rMutate = rMutate
        self.bit = bit
        self.qPool = queue.Queue(len)
        self.q = self.qPool.queue
        self.id = 1
        self.port = 10000
        self.pool = []
        self.bots = []
        self.typ = typ
        self.limit = limit

        self.first = True
        self.first2 = True

        self.power = 10
        self.w0 = wList[0]
        self.w1 = wList[1]
        self.w2 = wList[2]
        self.w3 = wList[3]

        #generator
        self.generate()

    def generate(self):
        for i in range(self.len):
            chrom = []
            for j in range(self.bit):
                input = randrange(0, 2)
                chrom.append(input)
            ch = Chrom(chrom, self.id, self.port)
            self.id = self.id + 1
            self.port = self.port + 1
            self.pool.append(ch)

    def add2q(self, chrom):
        if self.qPool.full():
            self.qPool.get()
        self.qPool.put(chrom)
    
    def getFront(self):
        return self.qPool.get()

    def getG(self, num):
        return self.q[num]

    # def initChrom(self):
    #     for i in self.pool:
    #         out = open("gene/bot" + str(i.getID()) + ".txt", "w")
    #         out.write(str(i.getGene()))
    #         out.close()

    def print(self):
        try:
            for i in range(self.len):
                print(self.getG(i))
        except:
            pass
    
    def printAll(self):
        try:
            for i in range(self.len):
                print(self.getG(i).getPair())
        except:
            pass
    
    def getPool(self):
        return self.pool

    def getResults(self, num):
        if self.typ == 1:
            openFile = open("botData/move/bot_" + str(num) + ".txt", "r").readlines()
        elif self.typ == 2:
            openFile = open("botData/atk/bot_" + str(num) + ".txt", "r").readlines()
        else:
            openFile = open("botData/def/bot_" + str(num) + ".txt", "r").readlines()
        temp = eval(openFile[0])
        return temp

    def calcFitAll(self, num):
        for bot in self.pool[:num]:
            stat = self.getResults(bot.getID())
            fitness = self.calcFitInd(stat)
            bot.setFit(fitness)
            self.add2q(bot)
        for bot in self.pool[:num]:
            self.queueOut(bot)
            self.pool.remove(bot)

    def calcFitInd(self, stat):
        spdScore = 0
        if self.typ == 1:
            for i in stat[4]:
                spdScore = spdScore + i
            fitness = self.w0 * stat[0] + spdScore - self.w3 * stat[3]
        elif self.typ == 2:
            fitness = self.w0 * stat[0] + self.w1 * stat[1]
        else:
            fitness = self.w0 * stat[0] - self.w2 * stat[2]
        return fitness

    def botComp(self):
        for i in self.pool:
            self.addBot(i)
    
    def addBot(self, chrom):
        b = Bot(chrom.getGene(), chrom.getID(), chrom.getPort(), self.typ, self.limit)
        self.bots.append(b)

    def test(self, num):
        tList = []
        for i in self.bots[:num]:
            t = Thread(target = i.run)
            t.start()
            tList.append(t)
            self.bots.remove(i)
        for t in tList:
            t.join()
        self.calcFitAll(num)
    
    def infiniTest(self, num):
        # for i in range(int(self.len/num)):
        while True:
            if len(self.bots) != 0:
                self.test(num)
            else:
                # break
                self.newKids(num)
                self.botComp()
                # self.initChrom()
            
    def newKids(self, num):
        tempFitList = []
        tempGenes = []
        wList = []
        rList = []
        for i in self.q:
            tempGenes.append(i.getGene())
            tempFitList.append(i.getFit())
        # print(tempFitList)
        for fit in tempFitList:
            if self.typ == 1:
                sFit = ((((100 ** (1/self.power)-1) * (fit-min(tempFitList)))/(max(tempFitList)-min(tempFitList)))+1) ** self.power
            elif self.typ == 2:
                sFit = fit 
            else:
                # sFit = fit
                sFit = ((((100 ** (1/self.power)-1) * (fit-min(tempFitList)))/(max(tempFitList)-min(tempFitList)))+1) ** self.power
            wList.append(sFit)
        for w in wList:
            relFit = (w/mean(wList))
            rList.append(relFit)
        self.wOut(rList)
        for i in range(num):
            parents = self.getParents(tempGenes, rList)
            child = self.genChild(parents)
            child = self.mutate(child)
            c = Chrom(child, self.id, self.port)
            self.id = self.id + 1
            self.port = self.port + 1
            self.pool.append(c)
    
    def wOut(self, lst):
        if self.first2:
            out = open("relFit.txt", "w")
            self.first2 = False
        else:
            out = open("relFit.txt", "a")
        out.write(str(lst) + "\n")
        out.close()

    def mutate(self, child):
        for i in range(self.bit):
            mut = (random() < self.rMutate)
            if mut:
                if child[i] == 0:
                    child[i] = 1
                else:
                    child[i] = 0
        return child

    def genChild(self, p):

        if self.typ == 1:
            p1 = p[randrange(0, 2)][0:6] 
            p2 = p[randrange(0, 2)][6:10] 
            p3 = p[randrange(0, 2)][10:19] 
            p4 = p[randrange(0, 2)][19:28]
            child = p1 + p2 + p3 + p4

        elif self.typ == 2:
            p1 = p[randrange(0, 2)][0:4]
            p2 = p[randrange(0, 2)][4:9]
            p3 = p[randrange(0, 2)][9:13]
            child = p1 + p2 + p3

        else:
            p1 = p[randrange(0, 2)][0:4]
            p2 = p[randrange(0, 2)][4:14]
            p3 = p[randrange(0, 2)][14:21]
            p4 = p[randrange(0, 2)][21:28]
            p5 = p[randrange(0, 2)][28:30]
            p6 = p[randrange(0, 2)][30:33]
            p7 = p[randrange(0, 2)][33:37]
            p8 = p[randrange(0, 2)][37:43]
            child = p1 + p2 + p3 + p4 + p5 + p6 + p7 + p8

        return child

    def getParents(self, gList, fList):
        parents = choices(gList, weights = fList, k = 2)
        return parents

    def getBotList(self):
        return self.bots
    
    def queueOut(self, obj):
        if self.typ == 1:
            if self.first:
                wb = Workbook()
                ws = wb.active
                wb.save(filename = "botData/bots_move_data.xlsx")
                out = open("botData/bots_move.txt", "w")
                self.first = False
            else:
                out = open("botData/bots_move.txt", "a")

            wb = openpyxl.load_workbook("botData/bots_move_data.xlsx")
            ws = wb.active

            lst = []
            for i in self.q:
                lst.append(i.getFit())

            ws.cell(column = 1, row = ws.max_row, value = ws.max_row - 1)
            ws.cell(column = 2, row = ws.max_row + 1, value = obj.getFit())
            
            
            wb.save("botData/bots_move_data.xlsx")
            out.write(str(obj.getPair()) + " " + str(mean(lst))  + "\n")
            out.close()
        elif self.typ == 2:
            if self.first:
                wb = Workbook()
                ws = wb.active
                wb.save(filename = "botData/bots_atk_data.xlsx")
                out = open("botData/bots_atk.txt", "w")
                self.first = False
            else:
                out = open("botData/bots_atk.txt", "a")

            wb = openpyxl.load_workbook("botData/bots_atk_data.xlsx")
            ws = wb.active

            lst = []
            for i in self.q:
                lst.append(i.getFit())

            ws.cell(column = 1, row = ws.max_row, value = ws.max_row - 1)
            ws.cell(column = 2, row = ws.max_row + 1, value = obj.getFit())
            
            
            wb.save("botData/bots_atk_data.xlsx")
            out.write(str(obj.getPair()) + " " + str(mean(lst))  + "\n")
            out.close()
        else:
            if self.first:
                wb = Workbook()
                ws = wb.active
                wb.save(filename = "botData/bots_def_data.xlsx")
                out = open("botData/bots_def.txt", "w")
                self.first = False
            else:
                out = open("botData/bots_def.txt", "a")

            wb = openpyxl.load_workbook("botData/bots_def_data.xlsx")
            ws = wb.active

            lst = []
            for i in self.q:
                lst.append(i.getFit())

            ws.cell(column = 1, row = ws.max_row, value = ws.max_row - 1)
            ws.cell(column = 2, row = ws.max_row + 1, value = obj.getFit())
            
            
            wb.save("botData/bots_def_data.xlsx")
            out.write(str(obj.getPair()) + " " + str(mean(lst))  + "\n")
            out.close()

def main():

    paraNum = 10 # Control the number of bots you want to parallely test at the same time

    #PARAMETER STRUCTURE
    # qga = QGA(LENGTH_CHROMOSOME, QUEUE_SIZE, MUTATION_RATE, [1 or 2 or 3], FITNESS_WEIGHTS, FRAME_LIMIT)
    # 1 = MOVE, 2 = ATTACK, 3 = DEFEND

    #MOVE TRAINING
    # qga = QGA(28, 250, 0.02, 1, [1, 0, 0, 0], 1680)

    #ATTACK TRAINING
    # qga = QGA(13, 250, 0.02, 2, [0, 200, 0, 0], 1680)

    #DEFEND TRAINING
    # qga = QGA(43, 250, 0.02, 3, [10, 0, 0, 0], 1680 * 3)

    qga.botComp()
    qga.infiniTest(paraNum)
    qga.printAll()

if __name__ == '__main__':
    main()